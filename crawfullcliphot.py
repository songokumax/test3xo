#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Crawl fullcliphot.org:
- Crawl listing pages, save thumbnail + title.
- Open each post with Playwright, capture network for .../note.txt URL.
- Download note.txt (M3U8-like), fetch all segment-*.png,
  extract m4s payload from PNG (chunk custom or trailing-after-IEND),
  then concat -> MP4 via ffmpeg.
- Grab post title & tags.
- Write rows to Excel.

Usage:
    python crawl_fullcliphot.py --start 1 --end 3 --out luuvideo --excel ketqua.xlsx
"""

import argparse
import os, re, sys, time, random, string, struct, subprocess, shutil, io, csv
from pathlib import Path
from urllib.parse import urljoin, urlparse
import requests
from bs4 import BeautifulSoup
import pandas as pd
from tqdm import tqdm
import logging
import sys

# ---------- Config ----------
BASE = "https://fullcliphot.org"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139 Safari/537.36"
}
CHUNK_TYPE = b"seGB"  # nếu site dùng chunk tuỳ chỉnh; nếu không có, script sẽ tự lấy phần sau IEND

FFMPEG_CMD = r"C:\ffmpeg-2025-10-27-git-68152978b5-full_build\bin\ffmpeg.exe"  # hoặc r"path\to\ffmpeg.exe" nếu chưa có trong PATH
# ---------------------------

# Cấu hình log file
LOG_FILE = "crawl_fullcliphot.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)

# Thay print() bằng logging.info / warning / error tuỳ loại
def log_info(msg): logging.info(msg)
def log_warn(msg): logging.warning(msg)
def log_err(msg): logging.error(msg)

def rand_name(min_len=20, max_len=40):
    n = random.randint(min_len, max_len)
    alphabet = string.ascii_letters + string.digits
    return "".join(random.choice(alphabet) for _ in range(n))

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def get_html(url):
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return r.text

def list_posts(listing_html):
    soup = BeautifulSoup(listing_html, "html.parser")
    posts = []
    for div in soup.select("#recent-content.content-loop .hentry"):
        a = div.select_one("a.thumbnail-link")
        img = div.select_one("a.thumbnail-link img")
        title_a = div.select_one(".entry-header h2.entry-title a")
        if not a or not img or not title_a:
            continue
        post_url = a["href"]
        thumb_url = img.get("src") or img.get("data-src") or ""
        title = title_a.get_text(strip=True)
        posts.append({"url": post_url, "thumb": thumb_url, "title": title})
    return posts

def save_thumbnail(thumb_url: str, out_dir: Path) -> Path:
    if not thumb_url:
        return Path()
    r = requests.get(thumb_url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    ext = os.path.splitext(urlparse(thumb_url).path)[1] or ".jpg"
    name = rand_name(10, 16) + ext
    out = out_dir / name
    with open(out, "wb") as f:
        f.write(r.content)
    return out

def extract_payload_from_png(png_bytes: bytes) -> bytes:
    # PNG signature
    if len(png_bytes) < 8 or png_bytes[:8] != b'\x89PNG\r\n\x1a\n':
        raise ValueError("Not a PNG")
    # scan chunks
    i = 8
    found_custom = None
    while i + 8 <= len(png_bytes):
        length = struct.unpack(">I", png_bytes[i:i+4])[0]; i += 4
        typ = png_bytes[i:i+4]; i += 4
        if i + length > len(png_bytes):
            break
        chunk_data = png_bytes[i:i+length]; i += length
        crc = png_bytes[i:i+4]; i += 4
        if typ == CHUNK_TYPE:
            found_custom = chunk_data
            break
        if typ == b"IEND":
            # trailing payload starts right after IEND CRC
            trailing = png_bytes[i:]
            if trailing:
                return trailing
    if found_custom is not None:
        return found_custom
    # fallback: try trailing after last IEND occurrence (if loop missed)
    idx = png_bytes.rfind(b"IEND")
    if idx != -1 and idx + 8 <= len(png_bytes):
        trailing = png_bytes[idx+8:]
        if trailing:
            return trailing
    raise RuntimeError("No payload (chunk or trailing) found in PNG")

def download_segments_from_playlist(playlist_text: str) -> list:
    # parse lines not starting with '#'
    urls = []
    for ln in playlist_text.splitlines():
        ln = ln.strip()
        if not ln or ln.startswith("#"):
            continue
        urls.append(ln)
    return urls

def fetch_playlist_url_with_playwright(playwright, post_url: str,
                                       timeout_ms=20000, retries=3,
                                       user_data_dir="chrome-profile") -> str:
    """
    Flow:
      - Mở post_url (fullcliphot.org), bắt /wp-admin/admin-ajax.php -> lấy 'data' (https://xfast.sbs/watch/....html)
      - Dùng Chrome thật (channel="chrome") + persistent profile để vào trang watch
      - Truyền Referer/Origin đúng khi goto để CDN/JWPlayer cấp playlist
      - Nghe network để bắt .../note.txt; nếu chưa thấy thì reload tối đa 'retries' lần
      - Fallback: regex note.txt trong HTML trang watch
    """
    import json, re, time, requests
    chromium = playwright.chromium

    # 1) Dùng Chrome thật + persistent profile (ổn định hơn headless Chromium)
    context = chromium.launch_persistent_context(
        user_data_dir=user_data_dir,                # tạo/tham chiếu thư mục profile
        channel="chrome",                           # chạy Chrome đã cài sẵn
        headless=True,                             # để debug trực quan; khi ổn có thể True
        args=[
            "--disable-blink-features=AutomationControlled",
            "--autoplay-policy=no-user-gesture-required",
        ],
    )
    # che navigator.webdriver + auto-mute khi play
    context.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        (function() {
          const _play = HTMLMediaElement.prototype.play;
          HTMLMediaElement.prototype.play = function() {
            try { this.muted = true; } catch(e) {}
            try { this.setAttribute && this.setAttribute('muted',''); } catch(e) {}
            return _play.call(this);
          };
        })();
    """)
    page = context.new_page()

    embed_url = None
    playlist_url = None
    capture_embed = True
    capture_note  = False

    def on_response(resp):
        nonlocal embed_url, playlist_url, capture_embed, capture_note
        u = resp.url
        # Bắt admin-ajax để lấy URL watch
        if capture_embed and "/wp-admin/admin-ajax.php" in u and "fullcliphot.org" in u:
            body = ""
            try: body = resp.text()
            except: pass
            try:
                j = json.loads(body)
                if isinstance(j, dict) and j.get("type") == "embed" and "data" in j:
                    cand = str(j["data"]).strip()
                    if cand.startswith("http"):
                        embed_url = cand
            except:
                m = re.search(r"https?://[^\s\"']+/watch/[^\s\"']+\.html", body or "")
                if m: embed_url = m.group(0)
            return

        # Bắt note.txt khi đã sang xfast.sbs
        if capture_note and "xfast.sbs" in u and "note.txt" in u:
            playlist_url = u

    page.on("response", on_response)

    # Helper: set extra headers (Referer/Origin) cho toàn bộ request trong context
    def set_ref_headers(ref):
        hdrs = {"Referer": ref}
        m = re.match(r"(https?://[^/]+)", ref)
        if m:
            hdrs["Origin"] = m.group(1)
        context.set_extra_http_headers(hdrs)

    try:
        # --- B1: vào trang bài để lấy embed_url ---
        page.goto(post_url, timeout=timeout_ms, wait_until="load")
        try: page.mouse.wheel(0, 1400)
        except: pass

        end = time.time() + timeout_ms/1000.0
        while time.time() < end and not embed_url:
            time.sleep(0.25)

        if not embed_url:
            html = page.content()
            m = re.search(r"https?://[^\s\"']+/watch/[^\s\"']+\.html", html)
            if m: embed_url = m.group(0)
        if not embed_url:
            return None

        # --- B2: sang trang watch bằng Chrome thật + đúng Referer/Origin ---
        capture_embed = False
        capture_note  = True

        # set headers cho toàn bộ subrequest (segments, note.txt…)
        set_ref_headers(post_url)

        for attempt in range(1, retries + 1):
            if attempt == 1:
                # Gắn referer trực tiếp cho main document
                page.goto(embed_url, timeout=timeout_ms, wait_until="load",
                          referer=post_url)
            else:
                page.reload(timeout=timeout_ms, wait_until="load")

            # cố gắng kích hoạt player
            try:
                page.mouse.wheel(0, 1400)
                page.click("css=.jw-display-icon-container, css=video", timeout=1500)
            except: pass
            try:
                page.evaluate("(window.jwplayer && jwplayer().play) && jwplayer().play();")
            except: pass

            end = time.time() + timeout_ms/1000.0
            while time.time() < end and not playlist_url:
                time.sleep(0.25)
            if playlist_url:
                break

        # --- B3: fallback regex trong HTML trang watch ---
        if not playlist_url:
            try:
                h = dict({"User-Agent": context.user_agent or HEADERS.get("User-Agent", "")})
                h["Referer"] = post_url
                m = re.match(r"(https?://[^/]+)", post_url)
                if m: h["Origin"] = m.group(1)
                whtml = requests.get(embed_url, headers=h, timeout=30).text
                m = re.search(r"https?://[^\s\"']+?/note\.txt", whtml)
                if m and "xfast.sbs" in m.group(0):
                    playlist_url = m.group(0)
            except: pass

        return playlist_url, embed_url   # <-- TRẢ VỀ CẢ HAI

    finally:
        context.close()

from pathlib import Path

def excel_append_row(xlsx_path, row):
    # row là dict: page, post_url, title, thumb_url, thumb_path, video_path, video_name, tags
    from openpyxl import Workbook, load_workbook
    cols = ["page","post_url","title","thumb_url","thumb_path","video_url","video_name","tags"]
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        wb = Workbook(); ws = wb.active
        ws.append(cols)
        wb.save(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    ws.append([row.get(c, "") for c in cols])
    wb.save(xlsx_path)

def ffmpeg_concat_m4s(seg_files: list, out_mp4: Path):
    # Kiểm tra ffmpeg có chạy được không
    import shutil
    if not shutil.which(FFMPEG_CMD):
        raise RuntimeError(
            f"Không tìm thấy FFmpeg tại '{FFMPEG_CMD}'. "
            f"Hãy sửa FFMPEG_CMD thành full path tới ffmpeg.exe hoặc thêm vào PATH."
        )

    # tạo list.txt (ASCII)
    list_txt = out_mp4.parent / (out_mp4.stem + "_list.txt")
    with open(list_txt, "w", encoding="ascii", newline="\n") as f:
        for p in seg_files:
            f.write(f"file '{str(p)}'\n")

    cmd = [
        FFMPEG_CMD, "-f", "concat", "-safe", "0",
        "-protocol_whitelist", "file,concat,crypto,data",
        "-i", str(list_txt),
        "-c", "copy", "-movflags", "+faststart",
        str(out_mp4)
    ]
    log_info(f"FFmpeg: {' '.join(cmd)}")
    try:
        p = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
    except FileNotFoundError:
        raise RuntimeError(
            f"Không chạy được FFmpeg (không tìm thấy '{FFMPEG_CMD}'). "
            f"Sửa FFMPEG_CMD hoặc PATH rồi chạy lại."
        )

    if p.returncode != 0:
        cmd2 = [
            FFMPEG_CMD, "-f", "concat", "-safe", "0",
            "-protocol_whitelist", "file,concat,crypto,data",
            "-i", str(list_txt),
            "-fflags", "+genpts", "-c", "copy", "-movflags", "+faststart",
            str(out_mp4)
        ]
        log_info(f"FFmpeg retry: {' '.join(cmd2)}")
        p2 = subprocess.run(cmd2, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        if p2.returncode != 0:
            raise RuntimeError("FFmpeg concat failed:\n" + p.stdout + "\n" + p2.stdout)


def get_post_meta(html: str):
    soup = BeautifulSoup(html, "html.parser")
    # title (trang bài)
    title = ""
    h1 = soup.select_one("h1.entry-title")
    if h1:
        title = h1.get_text(strip=True)
    # tags
    tags = [a.get_text(strip=True) for a in soup.select(".entry-tags .tag-links a")]
    return title, tags
    
def get_playlist_with_retry_only_on_timeout(pw, post_url: str, tries: int = 3, timeout_ms: int = 15000):
    """
    Chỉ retry khi fetch_playlist_url_with_playwright ném lỗi Page.goto timeout.
    Các lỗi khác: ném ra luôn (không retry).
    Trả về playlist_url (note.txt) hoặc ném lỗi sau khi thử đủ số lần.
    """
    last_timeout_err = None
    for i in range(tries):
        try:
            return fetch_playlist_url_with_playwright(pw, post_url, timeout_ms=timeout_ms)
        except PlaywrightTimeoutError as e:
            # chỉ bắt lỗi timeout của Playwright (đúng kiểu trong ảnh)
            log_info(f"   -> Page.goto timeout, thử lại {i+1}/{tries} ...")
            last_timeout_err = e
            time.sleep(2)
            continue
        except Exception:
            # lỗi khác: ném ra luôn để outer try/except xử lý như cũ
            raise
    # hết số lần retry mà vẫn timeout
    raise last_timeout_err if last_timeout_err else RuntimeError("Timeout nhưng không có exception?!")

from concurrent.futures import ThreadPoolExecutor, as_completed
from requests.adapters import HTTPAdapter, Retry

def build_session():
    sess = requests.Session()
    # retry nhẹ nhàng với backoff để tránh fail lẻ
    retries = Retry(total=3, backoff_factor=0.5, status_forcelist=(429, 500, 502, 503, 504))
    adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=retries)
    sess.mount("http://", adapter)
    sess.mount("https://", adapter)
    sess.headers.update(HEADERS)
    return sess

def _download_one(idx_url, sess: requests.Session, work_dir: Path):
    """Tải 1 PNG → rút payload → lưu đúng tên theo index. Trả về đường dẫn .m4s hoặc None."""
    idx, url = idx_url
    try:
        rb = sess.get(url, timeout=60).content
        payload = extract_payload_from_png(rb)
        out_seg = work_dir / f"seg_{idx:05d}.m4s"
        with open(out_seg, "wb") as f:
            f.write(payload)
        return out_seg
    except Exception as e:
        logging.warning(f"  - lỗi segment {idx}: {e}")
        return None

def parallel_download_segments(seg_urls: list, work_dir: Path, workers: int = 8) -> list[Path]:
    """
    Tải các segment PNG song song, nhưng ghi file theo đúng thứ tự chỉ số.
    Trả về danh sách Path đã được **sắp xếp theo index**.
    """
    ensure_dir(work_dir)
    sess = build_session()

    # Gắn index 1-based cho đúng thứ tự ghép
    jobs = [(i+1, u) for i, u in enumerate(seg_urls)]
    results = {}
    with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
        futures = {ex.submit(_download_one, j, sess, work_dir): j[0] for j in jobs}
        for fut in tqdm(as_completed(futures), total=len(futures), desc="Downloading PNG segments"):
            idx = futures[fut]
            p = fut.result()
            if p:
                results[idx] = p

    # Trả về theo thứ tự index
    ordered = [results[i] for i in sorted(results.keys())]
    return ordered

def crawl(args):
    out_root = Path(args.out).resolve()
    thumb_dir = out_root / "thumbs"
    video_dir = out_root / "videos"
    temp_dir = out_root / "tmp"
    ensure_dir(thumb_dir); ensure_dir(video_dir); ensure_dir(temp_dir)

    all_rows = []

    # Playwright lazy import (để không bắt buộc nếu chỉ test HTML)
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

    with sync_playwright() as pw:
        for page_no in range(args.start, args.end + 1):
            url = BASE if page_no == 1 else f"{BASE}/page/{page_no}/"
            os.system('cls' if os.name == 'nt' else 'clear')   # <== clear CMD khi sang page mới
            log_info(f"=== Listing page {page_no}: {url}")
            html = get_html(url)
            posts = list_posts(html)
            if not posts:
                log_warn("Không tìm thấy bài nào trên trang này.")
                continue

            for idx, post in enumerate(posts, 1):
                try:
                    log_info(f"\n[{page_no}.{idx}] {post['title']}")
                    # save thumb
                    thumb_path = save_thumbnail(post["thumb"], thumb_dir)
                    # lấy playlist (note.txt) qua network
                    playlist_url, video_url = None, ""
                    for retry in range(3):
                        try:
                            playlist_url, video_url = fetch_playlist_url_with_playwright(
                                pw, post["url"], timeout_ms=15000
                            )
                        except PlaywrightTimeoutError:
                            log_info(f"   -> Page.goto timeout, thử lại ({retry+1}/3)...")
                            time.sleep(2)
                            continue

                        if playlist_url:
                            break
                        log_warn(f"! Không tìm được note.txt → thử lại ({retry+1}/3)...")
                        time.sleep(3)

                    if not playlist_url:
                        log_warn("✗ Không tìm được note.txt sau 3 lần → bỏ qua bài này")
                        continue
                    log_info(f"note.txt = {playlist_url}")

                    # tải note.txt
                    playlist_text = requests.get(playlist_url, headers=HEADERS, timeout=30).text
                    seg_urls = download_segments_from_playlist(playlist_text)
                    if not seg_urls:
                        log_warn("! note.txt rỗng → bỏ qua")
                        continue

                    # tải từng PNG & tách payload
                    work_dir = temp_dir / rand_name(8, 10)
                    ensure_dir(work_dir)

                    # tải SONG SONG
                    seg_files = parallel_download_segments(seg_urls, work_dir, workers=args.seg_workers)

                    if not seg_files or len(seg_files) < len(seg_urls):
                        log_warn(f"! Tải thiếu segment ({len(seg_files)}/{len(seg_urls)}) – vẫn thử ghép")
                    if not seg_files:
                        log_warn("! Không có segment nào → bỏ qua")
                        shutil.rmtree(work_dir, ignore_errors=True)
                        continue

                    # ghép mp4
                    video_name = rand_name(20, 40) + ".mp4"
                    out_mp4 = video_dir / video_name
                    ffmpeg_concat_m4s(seg_files, out_mp4)

                    # metadata của bài
                    post_html = get_html(post["url"])
                    title, tags = get_post_meta(post_html)

                    # lưu row
                    row = {
                        "page": page_no,
                        "post_url": post["url"],
                        "title": title or post["title"],
                        "thumb_url": post["thumb"],
                        "thumb_path": os.path.basename(str(thumb_path)),  # chỉ tên file
                        "video_url": video_url or "",                    # thêm link watch từ admin-ajax
                        "video_name": video_name,
                        "tags": ", ".join(tags),
                    }
                    all_rows.append(row)
                    excel_append_row(args.excel, row)  # lưu nối sau mỗi phim
                    log_info(f"✓ DONE: {row['title']}")
                    log_info(f"   thumb: {row['thumb_path']}")
                    log_info(f"   file : {row['video_name']}")
                    log_info(f"   url  : {row['video_url']}")
                    log_info(f"   tags : {row['tags']}")

                    # dọn temp
                    shutil.rmtree(work_dir, ignore_errors=True)

                except Exception as e:
                    log_warn(f"!! Lỗi bài [{post['url']}]: {e}")

    if all_rows:
        # vẫn ghi hợp nhất nếu muốn (tuỳ chọn)
        df = pd.DataFrame(all_rows)
        excel_path = Path(args.excel).resolve()
        if excel_path.exists():
            old = pd.read_excel(excel_path)
            df = pd.concat([old, df], ignore_index=True)
        df.to_excel(excel_path, index=False)
        log_info(f"\n==> Đã lưu Excel hợp nhất: {excel_path}")
    else:
        # tạo file rỗng với header nếu chưa tồn tại
        from openpyxl import Workbook
        excel_path = Path(args.excel).resolve()
        if not excel_path.exists():
            wb = Workbook(); ws = wb.active
            ws.append(["page","post_url","title","thumb_url","thumb_path","video_url","video_name","tags"])
            wb.save(str(excel_path))
        log_warn(f"\n==> Không có bản ghi mới, file Excel: {excel_path}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--start", type=int, default=1)
    ap.add_argument("--end", type=int, default=3)
    ap.add_argument("--out", type=str, default="luuvideo")
    ap.add_argument("--excel", type=str, default="ketqua.xlsx")
    ap.add_argument("--seg_workers", type=int, default=8, help="Số luồng tải PNG segments")
    args = ap.parse_args()
    crawl(args)

if __name__ == "__main__":
    main()
